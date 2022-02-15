#This is the one with facebook_process_10_2021_detail_report.xlsx extenstion

import openpyxl
path= "facebook_process_10_2021_detail_report.xlsx"

wb_obj=openpyxl.load_workbook(path)

sheet_obj=wb_obj.active


row=sheet_obj.max_row
column=sheet_obj.max_column

print("total rows:", row)
print("total columns:", column)

for row in sheet_obj.iter_rows(min_row=1, min_col=1, max_row=60, max_col=17):
    for cell in row:
        print(cell.value, end="")
    print()


