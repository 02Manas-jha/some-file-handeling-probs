#This is the one with facebook_process_10_2021_detail_report.xlsx extenstion

import openpyxl
import csv
path= "facebook_process_10_2021_detail_report.xlsx"

wb_obj=openpyxl.load_workbook(path)

sheet_obj=wb_obj.active


row=sheet_obj.max_row
column=sheet_obj.max_column

print("total rows:", row)
print("total columns:", column)

required_cols = ['total', 'track_artist', 'song_name', 'isrcs', 'income', 'admin_Exp', 'royality']

for col in sheet_obj.iter_cols(min_row=1, min_col=1, max_row=60, max_col=17):
    # Iterating through required columns
    if col[1].value in required_cols:
        for cell in col:
            # None check validation
            if cell.value:
                print(cell.value)
        print('\n')


#This is facebook_process_10_2021_detail_report.csv extension
required_cols = ['total', 'track_artist', 'song_name', 'isrcs', 'income', 'admin_Exp', 'royality']
with open ('facebook_process_10_2021_detail_report.csv') as file:
    data = csv.DictReader(file)

    for row in data:
        for x in required_cols:
            print(row[x])
        print('\n')